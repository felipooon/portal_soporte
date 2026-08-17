from pathlib import Path
import xml.etree.ElementTree as ET
import zipfile
import openpyxl


def parsear_ods(ruta: Path) -> list[list[str]]:
    """Lee un archivo .ods (OpenDocument Spreadsheet) y retorna las filas como listas de strings."""
    try:
        with zipfile.ZipFile(ruta) as z:
            with z.open("content.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {
                    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
                    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
                }
                rows = []
                for row in root.findall(".//table:table-row", ns):
                    row_cells = []
                    for cell in row.findall(".//table:table-cell", ns):
                        repeat = int(cell.get("{urn:oasis:names:tc:opendocument:xmlns:table:1.0}number-columns-repeated", "1"))
                        text_nodes = cell.findall(".//text:p", ns)
                        cell_val = " ".join(t.text for t in text_nodes if t.text) if text_nodes else ""
                        row_cells.extend([cell_val] * min(repeat, 20))
                    if any(c for c in row_cells):
                        rows.append(row_cells)
                return rows
    except Exception as e:
        print(f"Error al parsear ODS: {e}")
        return []


def parsear_xlsx(ruta: Path) -> list[list[str]]:
    """Lee un archivo .xlsx (Excel) y retorna las filas como listas de strings."""
    try:
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
        ws = None
        for name in wb.sheetnames:
            if "alarm" in name.lower() or "config" in name.lower():
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active if wb.active else (wb.worksheets[0] if wb.worksheets else None)

        if not ws:
            return []
        rows = []
        for r in ws.iter_rows(values_only=True):
            r_str = [str(cell).strip() if cell is not None else "" for cell in r]
            if any(r_str):
                rows.append(r_str)
        return rows
    except Exception as e:
        print(f"Error al parsear XLSX: {e}")
        return []


def parsear_alarmas_excel(ruta_archivo: str | Path, nombre_centro: str | None = None) -> list[dict]:
    """
    Parsea una planilla de configuración de alarmas (.ods o .xlsx),
    omitiendo la columna 'Centro'.
    Retorna una lista de diccionarios representando cada fila de alarma.
    """
    ruta = Path(ruta_archivo)
    if not ruta.exists() or not ruta.is_file():
        return []

    ext = ruta.suffix.lower()
    if ext == ".ods":
        rows = parsear_ods(ruta)
    else:
        rows = parsear_xlsx(ruta)

    if not rows:
        return []

    def simplificar(texto: str) -> str:
        return str(texto or "").lower().replace(".", "").replace(" ", "").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")

    # Buscar fila de encabezados
    header_idx = 0
    norm_header_map: dict[str, int] = {}
    for idx, r in enumerate(rows):
        r_simples = [simplificar(cell) for cell in r]
        if any(k in r_simples for k in ["status", "equipo", "sensor", "medicion", "confmin"]):
            header_idx = idx
            for col_i, cell in enumerate(r):
                c_simp = simplificar(cell)
                if c_simp:
                    norm_header_map[c_simp] = col_i
            break

    def buscar_col(*posibles_nombres: str, default_idx: int) -> int:
        for nombre in posibles_nombres:
            s = simplificar(nombre)
            if s in norm_header_map:
                return norm_header_map[s]
        return default_idx

    idx_status = buscar_col("Status", default_idx=0)
    idx_centro = buscar_col("Centro", default_idx=1)
    idx_equipo = buscar_col("Equipo", default_idx=2)
    idx_sensor = buscar_col("Sensor", default_idx=3)
    idx_correo = buscar_col("Correo usuario", "Correo", "Usuario", default_idx=4)
    idx_min = buscar_col("Conf. Min.", "Conf Min", "Conf. Mínima", "Min", default_idx=5)
    idx_max = buscar_col("Conf. Max.", "Conf Max", "Conf. Máxima", "Max", default_idx=6)
    idx_medicion = buscar_col("Medición", "Medicion", "Variable", default_idx=7)
    idx_envio = buscar_col("Rango de envio (Minutos)", "Rango de envio", "Envio", default_idx=8 if len(rows[header_idx]) > 8 else -1)

    data_rows = rows[header_idx + 1:]

    def normalizar(s):
        return str(s or "").lower().replace(" ", "").replace("-", "").replace("_", "")

    target_norm = normalizar(nombre_centro) if nombre_centro else ""

    filas_filtradas = []
    filas_todas = []

    for r in data_rows:
        if not r or not any(r):
            continue

        centro_val = r[idx_centro] if idx_centro < len(r) else ""
        c_norm = normalizar(centro_val)

        status_val = r[idx_status] if idx_status < len(r) and r[idx_status] else "Activada"
        equipo_val = r[idx_equipo] if idx_equipo < len(r) and r[idx_equipo] else "-"
        sensor_val = r[idx_sensor] if idx_sensor < len(r) and r[idx_sensor] else "-"
        correo_val = r[idx_correo] if idx_correo < len(r) and r[idx_correo] else "-"
        min_val = r[idx_min] if idx_min < len(r) and r[idx_min] else "-"
        max_val = r[idx_max] if idx_max < len(r) and r[idx_max] else "-"
        med_val = r[idx_medicion] if idx_medicion < len(r) and r[idx_medicion] else "-"
        env_val = r[idx_envio] if 0 <= idx_envio < len(r) and r[idx_envio] else "60"

        item = {
            "status": status_val,
            "equipo": equipo_val,
            "sensor": sensor_val,
            "correo": correo_val,
            "conf_min": min_val,
            "conf_max": max_val,
            "medicion": med_val,
            "envio": env_val
        }

        filas_todas.append(item)
        if target_norm and (target_norm in c_norm or c_norm in target_norm):
            filas_filtradas.append(item)

    if filas_filtradas:
        return filas_filtradas

    return filas_todas


def limpiar_sensor_texto(texto_sensor: str) -> str:
    """
    Limpia cadenas largas de sensor provenientes del portal web/consola
    (ej: '(27943) Sensor 5 mts Pontón - Oxygen (oxygen) - Yatac (ac-yatac)')
    extrayendo únicamente la descripción limpia del sensor ('Sensor 5 mts Pontón').
    """
    import re
    if not texto_sensor or texto_sensor == "-":
        return "-"

    val = str(texto_sensor).strip()

    # 1. Si contiene guiones ' - ', la primera parte es la descripción del sensor
    if " - " in val:
        val = val.split(" - ")[0].strip()

    # 2. Eliminar el ID numérico entre paréntesis inicial, ej: '(27943) '
    val = re.sub(r"^\(\d+\)\s*", "", val).strip()

    return val if val else str(texto_sensor)


def normalizar_alarma_dict(item: dict) -> dict:
    """
    Normaliza y limpia un diccionario de alarma individual, detectando si 'equipo'
    trae la descripción de un sensor y corrigiendo el intercambio automáticamente.
    """
    import re
    status = str(item.get("status") or "Activo").strip()
    equipo = str(item.get("equipo") or "-").strip()
    sensor = str(item.get("sensor") or "-").strip()
    correo = str(item.get("correo") or "-").strip()
    conf_min = str(item.get("conf_min") or "-").strip()
    conf_max = str(item.get("conf_max") or "-").strip()
    medicion = str(item.get("medicion") or "-").strip()
    envio = str(item.get("envio") or "60").strip()

    # Intercambiar si 'equipo' trae la descripción del sensor
    es_sensor_en_equipo = (
        re.search(r"^\(\d+\)", equipo) or
        "sensor" in equipo.lower() or
        " - " in equipo or
        "pontón" in equipo.lower() or
        "ponton" in equipo.lower() or
        "jaula" in equipo.lower()
    )

    if es_sensor_en_equipo:
        # Extraer 'Equipo 1', 'Equipo 2', etc. si venía dentro
        match_eq = re.search(r"(Equipo\s*\d+)", equipo, re.IGNORECASE)
        eq_extraido = match_eq.group(1).title() if match_eq else "-"

        if sensor == "-" or not sensor or sensor.lower() == "sin sensor":
            sensor = equipo

        equipo = eq_extraido if eq_extraido != "-" else "Equipo 1"

    # Si 'equipo' aún trae paréntesis o IDs, limpiarlo
    if re.search(r"^\(\d+\)", equipo) or " - " in equipo:
        match_eq = re.search(r"(Equipo\s*\d+)", equipo, re.IGNORECASE)
        equipo = match_eq.group(1).title() if match_eq else "Equipo 1"

    sensor_clean = limpiar_sensor_texto(sensor)

    # Limpiar medicion
    if not medicion or medicion == "-" or re.search(r"^\(\d+\)", medicion) or " - " in medicion:
        s_low = (sensor + " " + equipo + " " + sensor_clean).lower()
        if "oxygen" in s_low or "oxigeno" in s_low or "oxígeno" in s_low or "oxi" in s_low:
            medicion = "Oxígeno"
        elif "salinity" in s_low or "salinidad" in s_low:
            medicion = "Salinidad"
        elif "temperature" in s_low or "temperatura" in s_low:
            medicion = "Temperatura"
        elif "orp" in s_low:
            medicion = "ORP"
        elif "ph" in s_low:
            medicion = "pH"
        elif "conductivid" in s_low or "conductivity" in s_low:
            medicion = "Conductividad"
        else:
            medicion = "Oxígeno"

    return {
        "status": status if status else "Activo",
        "equipo": equipo if equipo else "-",
        "sensor": sensor_clean if sensor_clean else "-",
        "correo": correo if correo else "-",
        "conf_min": conf_min if conf_min else "-",
        "conf_max": conf_max if conf_max else "-",
        "medicion": medicion if medicion else "Oxígeno",
        "envio": envio if envio else "60"
    }


def parsear_alarmas_texto(texto: str, nombre_centro: str | None = None) -> list[dict]:
    """
    Parsea texto copiado de tablas de alarmas (Excel/Web/Consola), soportando columnas
    separadas por TAB (\t), punto y coma (;), coma (,) o espacio múltiple.
    """
    import re
    if not texto or not texto.strip():
        return []

    lineas = [l.strip() for l in texto.strip().splitlines() if l.strip()]
    if not lineas:
        return []

    def simplificar(t: str) -> str:
        return str(t or "").lower().replace(".", "").replace(" ", "").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")

    primera_linea = lineas[0]
    if "\t" in primera_linea:
        sep = "\t"
    elif ";" in primera_linea:
        sep = ";"
    elif "," in primera_linea and primera_linea.count(",") > 3:
        sep = ","
    else:
        sep = None

    def dividir_linea(lin: str) -> list[str]:
        if sep:
            return [c.strip() for c in lin.split(sep)]
        return [c.strip() for c in re.split(r"\t+|\s{2,}", lin) if c.strip()]

    matriz = [dividir_linea(l) for l in lineas]
    if not matriz:
        return []

    header_idx = -1
    norm_header_map: dict[str, int] = {}
    r_simples_0 = [simplificar(c) for c in matriz[0]]

    if any(k in r_simples_0 for k in ["status", "estado", "equipo", "sensor", "medicion", "usuario", "minima", "maxima"]):
        header_idx = 0
        for col_i, cell in enumerate(matriz[0]):
            c_simp = simplificar(cell)
            if c_simp:
                norm_header_map[c_simp] = col_i

    def buscar_col(*posibles_nombres: str, default_idx: int) -> int:
        for nombre in posibles_nombres:
            s = simplificar(nombre)
            if s in norm_header_map:
                return norm_header_map[s]
        return default_idx

    if header_idx == 0:
        idx_status = buscar_col("Estado", "Status", default_idx=0)
        idx_correo = buscar_col("Usuario", "Correo", "Correo usuario", default_idx=1)
        idx_min = buscar_col("Mínima", "Minima", "Conf. Min.", "Min", default_idx=2)
        idx_max = buscar_col("Máxima", "Maxima", "Conf. Max.", "Max", default_idx=3)
        idx_medicion = buscar_col("Medicion Especifica", "Medicion", "Medición", "Variable", default_idx=4)
        idx_centro = buscar_col("Centros", "Centro", default_idx=5)
        idx_equipo = buscar_col("Equipo", default_idx=6)
        idx_sensor = buscar_col("Sensor", default_idx=7)
        data_rows = matriz[1:]
    else:
        estados_validos = {"activo", "activa", "activada", "inactivo", "inactiva", "desactivada", "ok", "enabled", "disabled"}
        filas_validas = [r for r in matriz if len(r) >= 4 and simplificar(r[0]) in estados_validos]
        if not filas_validas:
            return []
        idx_status = 0
        idx_correo = 1
        idx_min = 2
        idx_max = 3
        idx_medicion = 4
        idx_centro = 5
        idx_equipo = 6
        idx_sensor = 7
        data_rows = filas_validas

    alarmas = []
    for r in data_rows:
        if not r or not any(r):
            continue

        if header_idx == 0 and r == matriz[0]:
            continue
        if simplificar(r[0]) in ("estado", "status") and len(r) > 1 and simplificar(r[1]) in ("usuario", "correo"):
            continue

        status_val = r[idx_status] if 0 <= idx_status < len(r) and r[idx_status] else "Activo"
        correo_val = r[idx_correo] if 0 <= idx_correo < len(r) and r[idx_correo] else "-"
        min_val = r[idx_min] if 0 <= idx_min < len(r) and r[idx_min] else "-"
        max_val = r[idx_max] if 0 <= idx_max < len(r) and r[idx_max] else "-"
        med_val = r[idx_medicion] if 0 <= idx_medicion < len(r) and r[idx_medicion] else ""
        equipo_val = r[idx_equipo] if 0 <= idx_equipo < len(r) and r[idx_equipo] else "-"
        sensor_raw = r[idx_sensor] if 0 <= idx_sensor < len(r) and r[idx_sensor] else "-"

        item_raw = {
            "status": status_val,
            "equipo": equipo_val,
            "sensor": sensor_raw,
            "correo": correo_val,
            "conf_min": min_val,
            "conf_max": max_val,
            "medicion": med_val,
            "envio": "60"
        }

        alarmas.append(normalizar_alarma_dict(item_raw))

    return alarmas
